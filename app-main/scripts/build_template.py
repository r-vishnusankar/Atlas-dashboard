"""Generate /app/frontend/public/stresk_template.xlsx (and matching CSV) so the
Stresk dashboard ships with a ready-to-use Excel / Google Sheet template.

Columns mirror the schema consumed by js/lib/helpers.js::csvToProjects.
Includes:
  * Styled header row (Google Blue)
  * Data validation (dropdowns) on stage / status / priority
  * Frozen header row, sensible column widths
  * 32 pre-filled sample projects (same deterministic data as the dashboard mock)
"""

from __future__ import annotations

import csv
import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = Path("/app/frontend/public")
XLSX_PATH = OUT_DIR / "stresk_template.xlsx"
CSV_PATH = OUT_DIR / "stresk_sheet_template.csv"

HEADERS = [
    "project_id",
    "project_name",
    "owner",
    "stage",
    "status",
    "progress",
    "start_date",
    "release_date",
    "priority",
    "client",
    "tags",
    "notes",
]

STAGES = ["Planning", "Development", "QA", "Release", "Live"]
STATUSES = ["on_track", "at_risk", "delayed"]
PRIORITIES = ["High", "Medium", "Low"]

PROJECT_NAMES = [
    "Acme Corp Website",
    "Nimbus Health Patient Portal",
    "Orbital Labs Admin Dashboard",
    "Ironclad Finance Mobile App",
    "Lumina Retail Checkout 2.0",
    "Vertex Analytics Pipeline",
    "Polaris Media CMS Migration",
    "Glider Logistics Tracker",
    "Aster Robotics Control Panel",
    "Brightwave Edu LMS",
    "Cascade Energy Billing Rewrite",
    "Delphi Legal Contract AI",
    "Helios Solar Monitor",
    "Kestrel HR Onboarding",
    "Lotus Hospitality Booking",
    "Meridian Bank Open API",
    "Northwind Inventory Sync",
    "Onyx Insurance Claims Portal",
    "Quartz Gaming Leaderboards",
    "Rivet Manufacturing IoT",
    "Stellar Telecom Self-Service",
    "Tidepool Fitness App",
    "Umbra Security Audit Tool",
    "Vanta Compliance Sync",
    "Willow Gardens Store",
    "Xenon Travel Itineraries",
    "Yonder Social Feed",
    "Zephyr Auto Service Hub",
    "Cobalt Studio Portfolio",
    "Ember Podcast Platform",
    "Fable Reader App",
    "Gemstone Marketplace",
]

CLIENTS = [
    "Acme Corp", "Nimbus Health", "Orbital Labs", "Ironclad Finance",
    "Lumina Retail", "Vertex Analytics", "Polaris Media", "Glider Logistics",
    "Aster Robotics", "Brightwave Edu", "Cascade Energy", "Delphi Legal",
    "Helios Solar", "Kestrel HR", "Lotus Hospitality", "Meridian Bank",
    "Northwind Co.", "Onyx Insurance", "Quartz Gaming", "Rivet Mfg.",
    "Stellar Telecom", "Tidepool Inc.", "Umbra Security", "Vanta",
    "Willow Gardens", "Xenon Travel", "Yonder", "Zephyr Auto",
    "Cobalt Studio", "Ember Media", "Fable", "Gemstone",
]

OWNERS = [
    "Sarah K.", "Marcus T.", "Priya R.", "Diego N.", "Aya M.",
    "Liam O.", "Zoe F.", "Ravi S.", "Nora W.", "Elif D.",
    "Kenji H.", "Maya P.",
]

TAGS = [
    "E-commerce", "React", "AI", "Mobile", "API", "Dashboard",
    "Internal", "GDPR", "Payments", "Analytics", "SEO", "Migration",
    "Onboarding", "Multi-tenant", "Realtime", "B2B", "B2C",
]

NOTES = [
    "Waiting on content from client marketing team.",
    "Third-party API rate limits being reviewed.",
    "Security review scheduled for next week.",
    "Designs signed off — implementation in progress.",
    "Dependency upgrade required before ship.",
    "Smoke tests passing on staging.",
    "Client requested scope additions.",
    "Load testing in progress.",
    "",
    "Awaiting stakeholder approval.",
    "Roll-forward migration plan drafted.",
    "Localization pending for 3 markets.",
]

HERO_PROFILES = [
    ("Development", "delayed", -12, "High"),
    ("QA", "at_risk", 3, "High"),
    ("Development", "at_risk", 6, "Medium"),
    ("Planning", "delayed", -5, "High"),
    ("Release", "on_track", 9, "High"),
    ("Live", "on_track", -30, "Low"),
    ("Live", "on_track", -65, "Medium"),
    ("Development", "on_track", 42, "Medium"),
    ("Development", "on_track", 60, "High"),
    ("QA", "on_track", 14, "Medium"),
]

# Deterministic — mirrors mulberry32(7) style variety without exact parity.
RNG = random.Random(7)

STAGE_PROGRESS = {
    "Planning": (2, 20),
    "Development": (20, 65),
    "QA": (55, 85),
    "Release": (75, 98),
    "Live": (100, 100),
}


def build_rows() -> list[dict]:
    today = dt.date.today()
    rows: list[dict] = []
    for i, name in enumerate(PROJECT_NAMES):
        if i < len(HERO_PROFILES):
            stage, status, offset, priority = HERO_PROFILES[i]
        else:
            stage = RNG.choice(STAGES)
            status = RNG.choice(STATUSES)
            offset = RNG.randint(-45, 105)
            priority = RNG.choice(PRIORITIES)

        release = today + dt.timedelta(
            days=min(offset, -14) if stage == "Live" else offset
        )
        start = release - dt.timedelta(days=RNG.randint(30, 120))

        lo, hi = STAGE_PROGRESS[stage]
        progress = RNG.randint(lo, hi)
        if status == "delayed" and stage != "Live":
            progress = max(5, progress - 20)

        rows.append({
            "project_id": f"PRJ-{i + 1:03d}",
            "project_name": name,
            "owner": RNG.choice(OWNERS),
            "stage": stage,
            "status": status,
            "progress": progress,
            "start_date": start.isoformat(),
            "release_date": release.isoformat(),
            "priority": priority,
            "client": CLIENTS[i] if i < len(CLIENTS) else RNG.choice(CLIENTS),
            "tags": ", ".join(RNG.sample(TAGS, RNG.randint(1, 3))),
            "notes": RNG.choice(NOTES),
        })
    return rows


def write_csv(rows: list[dict]) -> None:
    with CSV_PATH.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=HEADERS)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def write_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    # --- Header row ---------------------------------------------------------
    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(name="Inter", bold=True, size=11, color="FFFFFF")
    header_align = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # --- Data rows ----------------------------------------------------------
    body_font = Font(name="Inter", size=11, color="202124")
    body_align = Alignment(vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(HEADERS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[key])
            cell.font = body_font
            cell.alignment = body_align
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8F9FA")

    # --- Dropdown data validations -----------------------------------------
    max_row = len(rows) + 200  # extend validations so users can add rows
    validations = {
        "D": ",".join(STAGES),      # stage
        "E": ",".join(STATUSES),    # status
        "I": ",".join(PRIORITIES),  # priority
    }
    for col_letter, options in validations.items():
        dv = DataValidation(
            type="list",
            formula1=f'"{options}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f"Please pick one of: {options}",
        )
        dv.add(f"{col_letter}2:{col_letter}{max_row}")
        ws.add_data_validation(dv)

    # Progress (column F) — integer 0-100
    dv_progress = DataValidation(
        type="whole",
        operator="between",
        formula1=0,
        formula2=100,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Out of range",
        error="Progress must be an integer between 0 and 100.",
    )
    dv_progress.add(f"F2:F{max_row}")
    ws.add_data_validation(dv_progress)

    # Date columns (G, H) — date values only
    for col_letter in ("G", "H"):
        dv_date = DataValidation(
            type="date",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid date",
            error="Use a YYYY-MM-DD style date.",
        )
        dv_date.add(f"{col_letter}2:{col_letter}{max_row}")
        ws.add_data_validation(dv_date)

    # --- Column widths ------------------------------------------------------
    widths = {
        "A": 11, "B": 34, "C": 16, "D": 14, "E": 12,
        "F": 11, "G": 13, "H": 13, "I": 11, "J": 22,
        "K": 28, "L": 42,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    # --- Second sheet: README ----------------------------------------------
    readme = wb.create_sheet("README")
    readme_lines = [
        ("Stresk Dashboard — Data Template", True),
        ("", False),
        ("How to publish this sheet to your dashboard:", True),
        ("1. Upload this file to Google Drive and open with Google Sheets.", False),
        ("2. File → Share → Publish to web.", False),
        ("3. Choose the 'Projects' tab, format = Comma-separated values (.csv).", False),
        ("4. Click Publish and copy the generated URL.", False),
        ("5. In the dashboard repo, open src/config.js and paste the URL into CONFIG.SHEET_URL.", False),
        ("6. Reload the dashboard — the 'mock' pill in the header switches to 'sheet'.", False),
        ("", False),
        ("Column rules:", True),
        ("• project_id — any unique string (e.g. PRJ-001).", False),
        ("• stage — one of: Planning, Development, QA, Release, Live.", False),
        ("• status — one of: on_track, at_risk, delayed.", False),
        ("• progress — integer 0–100.", False),
        ("• start_date / release_date — use YYYY-MM-DD.", False),
        ("• priority — one of: High, Medium, Low.", False),
        ("• tags — comma-separated list inside a single cell.", False),
        ("", False),
        ("The dashboard auto-computes OVERDUE / AT-RISK / STALLED alerts from these fields.", False),
    ]
    for r_idx, (text, bold) in enumerate(readme_lines, start=1):
        cell = readme.cell(row=r_idx, column=1, value=text)
        cell.font = Font(name="Inter", bold=bold, size=13 if bold else 11, color="202124")
        cell.alignment = Alignment(vertical="center", wrap_text=False)
    readme.column_dimensions["A"].width = 110
    readme.sheet_view.showGridLines = False

    wb.save(XLSX_PATH)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    print(f"Wrote {CSV_PATH} ({len(rows)} rows)")
    print(f"Wrote {XLSX_PATH} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
